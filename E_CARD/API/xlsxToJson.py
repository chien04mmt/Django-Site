from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd

from artice_car.settings import MEDIA_ROOT


def XLSX_TO_JSON(column,sql):
    with open('CSV/tieuthu.csv','w',encoding='utf-8') as w:
        w.write(column+'\n')
        
        for item in sql:
            chuoi = str(item)
            w.write(chuoi[1:-1].replace("'",'').replace(",",'!').replace("! ",',').replace("!000",'').replace(".00",'').replace("None",'')+'\n')   
                       
    df = pd.read_csv('CSV/tieuthu.csv',encoding='utf-8')
    df.to_excel('CSV/output.xlsx', 'tieuthu')

    # import excel2json
    # excel2json.convert_from_file('CSV/output.xlsx')


    wb = load_workbook(filename='CSV/output.xlsx')
    ws = wb.active

    my_list = []

    last_column = len(list(ws.columns))
    last_row = len(list(ws.rows))

    for row in range(1, last_row + 1):
        my_dict = {}
        for column in range(1, last_column + 1):
            column_letter = get_column_letter(column)
            if row > 1:
                my_dict[ws[column_letter + str(1)].value] = ws[column_letter + str(row)].value
        my_list.append(my_dict)
    # print(my_list)
    return my_list
    # data = json.dumps(my_list, sort_keys=False, indent=4)
    # with open('CSV/output.json', 'w', encoding='utf-8') as f:
    #     f.write(data)



def XLSX_TO_JSON2(filename):

    wb = load_workbook(filename= MEDIA_ROOT+'/'+filename)
    ws = wb.active

    my_list = []

    last_column = len(list(ws.columns))
    last_row = len(list(ws.rows))

    for row in range(1, last_row + 1):
        my_dict = {}
        for column in range(1, last_column + 1):
            column_letter = get_column_letter(column)           
            if row > 1:
                value= ws[column_letter + str(row)].value
                my_dict[ws[column_letter + str(1)].value] = (value if value is not None else '')#Chuyển các giá trị None về ''
        if len(my_dict)>0: my_list.append(my_dict)#Bỏ qua các oject list bị trống
    return my_list
